"""Exam blueprint: read the authored Full_Exam workbook and turn one sheet's
question distribution into a list of generation tasks the existing pipeline can
run.

The authored sheets (Full_Exam/Qudrat_Arabic.xlsx) use an ARABIC taxonomy
(القسم / التصنيف / التصنيف الفرعي). The generation pipeline (app6_final via
generate_visual) is driven by the ENGLISH taxonomy of the corpus workbook
(Qudrat Sample Sheet _ v2 (1).xlsx). This module maps Arabic -> English so each
blueprint row becomes a (section, category, subcat) the pipeline understands and
that actually has corpus examples (so app6 never hits "No example questions").

Canonical full exam = Khaled-Sample_1_Arabic: 120 questions (61 quant + 59 verbal).
"""

from collections import Counter, OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import openpyxl

ROOT = Path(__file__).resolve().parent
BLUEPRINT_XLSX = ROOT / "Full_Exam" / "Qudrat_Arabic.xlsx"
DEFAULT_SHEET = "Khaled-Sample_1_Arabic"

# Column indices in the authored sheets (0-based), confirmed from the workbook.
COL_QNO, COL_Q = 0, 1
COL_SECTION, COL_CATEGORY, COL_SUBCAT = 8, 9, 10

# Section: Arabic -> pipeline section name.
SECTION_MAP = {"كمي": "Quantitative", "لفظي": "Verbal"}

# Quant sub-category (Arabic) -> (English Category, English Sub-Category).
# Sub-Category strings are chosen to EXIST in the corpus so example filtering
# never returns empty; "ANY" means generate at category level (no subcat filter).
QUANT_MAP = {
    # الحساب bucket -> Arithmetic
    "الأعداد وخواصها": ("Arithmetic", "ANY"),
    "الأنماط والمتتابعات": ("Arithmetic", "Sequences & Series"),
    "الجذور": ("Arithmetic", "Powers & Roots"),
    "القواسم والمضاعفات والتحليل والوحدات": ("Arithmetic", "ANY"),
    "الكسور الاعتيادية": ("Arithmetic", "Fractions & Decimals"),
    "الكسور العشرية": ("Arithmetic", "Fractions & Decimals"),
    "النسبة المئوية": ("Arithmetic", "Percentages"),
    "النسبة والتناسب": ("Arithmetic", "Ratios & Proportions"),
    "الأسس": ("Arithmetic", "Powers & Roots"),
    # الهندسة bucket -> Geometry (figure-bearing; routed to geo kernel)
    "الأشكال الرباعية": ("Geometry", "Quadrilaterals & Polygons"),
    "الدائرة": ("Geometry", "Circles (Arcs, Chords, Tangents)"),
    "المثلثات": ("Geometry", "Angles & Triangles"),
    "المساحات المظللة": ("Geometry", "Area, Perimeter, Volume, Surface Area"),
    "المستقيمات والزوايا": ("Geometry", "Angles & Triangles"),
    "الهندسة الإحداثية + إستراتيجيات الحل السريع": ("Geometry", "Coordinate Geometry"),
    # تفسير البيانات -> Data Interpretation (tables/charts -> matplotlib figure)
    "الإحصاء": ("Data Interpretation & Reasoning", "Tables/Charts/Graphs"),
}

# Verbal sub-category (Arabic) -> (English Category, English Sub-Category).
VERBAL_MAP = {
    "تناظر لفظي": ("Analogies & Word Relationships", "ANY"),
    "إكمال جمل": ("Sentence Completion", "ANY"),
    "استيعاب المقروء": ("Reading Comprehension", "ANY"),
    # No clean corpus category for these two minor types; map to nearest.
    "خطأ سياقي": ("Reading Comprehension", "ANY"),
    "المفردة الشاذة": ("Analogies & Word Relationships", "ANY"),
}


@dataclass
class GenTask:
    section: str          # "Quantitative" | "Verbal"
    category: str         # English corpus category
    subcat: str           # English corpus sub-category or "ANY"
    count: int            # how many questions to generate
    lang: str             # "Arabic" | "English"
    ar_section: str       # original Arabic section (for display/grouping)
    ar_subcat: str        # original Arabic sub-category (for display)

    def key(self):
        return (self.section, self.category, self.subcat)


def _map_row(ar_section: str, ar_cat: str, ar_sub: str):
    """Return (section, category, subcat) in pipeline terms, or None to skip."""
    section = SECTION_MAP.get((ar_section or "").strip())
    if not section:
        return None
    sub = (ar_sub or "").strip()
    table = QUANT_MAP if section == "Quantitative" else VERBAL_MAP
    mapped = table.get(sub)
    if mapped:
        return section, mapped[0], mapped[1]
    # Unknown sub-category: fall back to category level so generation still runs.
    return section, ("Arithmetic" if section == "Quantitative"
                     else "Sentence Completion"), "ANY"


def load_blueprint(sheet: str = DEFAULT_SHEET, lang: str = "Arabic",
                   xlsx: Path = BLUEPRINT_XLSX) -> List[GenTask]:
    """Parse one sheet into a merged list of generation tasks.

    Rows mapping to the same (section, category, subcat) are merged and their
    counts summed, so the pipeline is called once per distinct task.
    """
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not in {xlsx.name}: {wb.sheetnames}")
    ws = wb[sheet]

    counts: "OrderedDict[tuple, dict]" = OrderedDict()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if not row or row[COL_Q] in (None, "", 0, "0"):
            continue
        mapped = _map_row(row[COL_SECTION], row[COL_CATEGORY], row[COL_SUBCAT])
        if not mapped:
            continue
        ar_section = (row[COL_SECTION] or "").strip()
        ar_sub = (row[COL_SUBCAT] or "").strip()
        key = mapped
        if key not in counts:
            counts[key] = {"count": 0, "ar_section": ar_section, "ar_sub": ar_sub}
        counts[key]["count"] += 1

    tasks = [
        GenTask(section=k[0], category=k[1], subcat=k[2], count=v["count"],
                lang=lang, ar_section=v["ar_section"], ar_subcat=v["ar_sub"])
        for k, v in counts.items()
    ]
    return tasks


def summarize(tasks: List[GenTask]) -> dict:
    total = sum(t.count for t in tasks)
    by_section = Counter()
    for t in tasks:
        by_section[t.section] += t.count
    return {"total": total, "by_section": dict(by_section), "n_tasks": len(tasks)}


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Inspect an exam blueprint")
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--lang", default="Arabic")
    args = ap.parse_args()
    tasks = load_blueprint(args.sheet, args.lang)
    print(summarize(tasks))
    for t in tasks:
        print(f"  {t.count:>3}  {t.section:<13} {t.category} / {t.subcat}"
              f"   [{t.ar_subcat}]")
